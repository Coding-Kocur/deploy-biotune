import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy

universities = [
    {
        "name": "Technische Universität Wien",
        "country": "Austria",
        "city": "Vienna",
        "qs_num": 192,
        "qs": "~192",
        "students": 28000,
        "temp": 3.6,
        "col_index": 73,
        "website": "https://www.tuwien.at/en/",
        "maps": "https://www.google.com/maps/place/TU+Wien,+Vienna,+Austria",
    },
    {
        "name": "Tampere University",
        "country": "Finlandia",
        "city": "Tampere",
        "qs_num": 382,
        "qs": "~382",
        "students": 21000,
        "temp": -3.0,
        "col_index": 72,
        "website": "https://www.tuni.fi/en",
        "maps": "https://www.google.com/maps/place/Tampere+University,+Tampere,+Finland",
    },
    {
        "name": "University of Pisa",
        "country": "Włochy",
        "city": "Pisa",
        "qs_num": 392,
        "qs": "~392",
        "students": 50000,
        "temp": 9.6,
        "col_index": 62,
        "website": "https://www.unipi.it/index.php/english",
        "maps": "https://www.google.com/maps/place/University+of+Pisa,+Pisa,+Italy",
    },
    {
        "name": "Aristotle University of Thessaloniki",
        "country": "Grecja",
        "city": "Thessaloniki",
        "qs_num": 521,
        "qs": "~521",
        "students": 73000,
        "temp": 9.2,
        "col_index": 48,
        "website": "https://www.auth.gr/en/",
        "maps": "https://www.google.com/maps/place/Aristotle+University+of+Thessaloniki,+Greece",
    },
    {
        "name": "Università Cattolica del Sacro Cuore",
        "country": "Włochy",
        "city": "Milan",
        "qs_num": 526,
        "qs": "~526",
        "students": 37000,
        "temp": 6.0,
        "col_index": 72,
        "website": "https://www.unicatt.it/",
        "maps": "https://www.google.com/maps/place/Università+Cattolica+del+Sacro+Cuore,+Milan,+Italy",
    },
    {
        "name": "University of Cyprus",
        "country": "Cypr",
        "city": "Nicosia",
        "qs_num": 571,
        "qs": "~571",
        "students": 7000,
        "temp": 14.2,
        "col_index": 58,
        "website": "https://www.ucy.ac.cy/en/",
        "maps": "https://www.google.com/maps/place/University+of+Cyprus,+Nicosia",
    },
    {
        "name": "TalTech – Tallinn University of Technology",
        "country": "Estonia",
        "city": "Tallinn",
        "qs_num": 651,
        "qs": "~651",
        "students": 10000,
        "temp": -0.2,
        "col_index": 52,
        "website": "https://taltech.ee/en/",
        "maps": "https://www.google.com/maps/place/Tallinn+University+of+Technology,+Tallinn,+Estonia",
    },
    {
        "name": "Universidade de Aveiro",
        "country": "Portugalia",
        "city": "Aveiro",
        "qs_num": 681,
        "qs": "~681",
        "students": 15000,
        "temp": 12.0,
        "col_index": 45,
        "website": "https://www.ua.pt/",
        "maps": "https://www.google.com/maps/place/University+of+Aveiro,+Aveiro,+Portugal",
    },
    {
        "name": "University of Malta",
        "country": "Malta",
        "city": "Msida",
        "qs_num": 801,
        "qs": "~801",
        "students": 12000,
        "temp": 15.2,
        "col_index": 60,
        "website": "https://www.um.edu.mt/",
        "maps": "https://www.google.com/maps/place/University+of+Malta,+Msida,+Malta",
    },
    {
        "name": "Università degli Studi di Cagliari",
        "country": "Włochy",
        "city": "Cagliari",
        "qs_num": 851,
        "qs": "~851",
        "students": 27000,
        "temp": 12.2,
        "col_index": 55,
        "website": "https://www.unica.it/",
        "maps": "https://www.google.com/maps/place/University+of+Cagliari,+Cagliari,+Italy",
    },
    {
        "name": "Technical University of Liberec",
        "country": "Czechy",
        "city": "Liberec",
        "qs_num": 1001,
        "qs": "~1001",
        "students": 7000,
        "temp": 1.6,
        "col_index": 42,
        "website": "https://www.tul.cz/en/",
        "maps": "https://www.google.com/maps/place/Technical+University+of+Liberec,+Czechia",
    },
    {
        "name": "Balıkesir Üniversitesi",
        "country": "Turcja",
        "city": "Balıkesir",
        "qs_num": 1201,
        "qs": "~1201",
        "students": 45000,
        "temp": 8.0,
        "col_index": 27,
        "website": "https://www.balikesir.edu.tr/",
        "maps": "https://www.google.com/maps/place/Balıkesir+University,+Balıkesir,+Turkey",
    },
    {
        "name": "Varna Free University",
        "country": "Bułgaria",
        "city": "Varna",
        "qs_num": 1401,
        "qs": "~1401",
        "students": 8000,
        "temp": 6.0,
        "col_index": 33,
        "website": "https://www.vfu.bg/en/",
        "maps": "https://www.google.com/maps/place/Varna+Free+University,+Varna,+Bulgaria",
    },
    {
        "name": "Instituto Politécnico do Porto",
        "country": "Portugalia",
        "city": "Porto",
        "qs_num": 9000,
        "qs": "b/r",
        "students": 18000,
        "temp": 11.4,
        "col_index": 45,
        "website": "https://www.ipp.pt/",
        "maps": "https://www.google.com/maps/place/Polytechnic+Institute+of+Porto,+Portugal",
    },
    {
        "name": "TED University",
        "country": "Turcja",
        "city": "Ankara",
        "qs_num": 9001,
        "qs": "b/r",
        "students": 4500,
        "temp": 4.2,
        "col_index": 27,
        "website": "https://www.tedu.edu.tr/en",
        "maps": "https://www.google.com/maps/place/TED+University,+Ankara,+Turkey",
    },
    {
        "name": "Windesheim University of Applied Sciences",
        "country": "Niderlandy",
        "city": "Zwolle",
        "qs_num": 9002,
        "qs": "b/r",
        "students": 27000,
        "temp": 4.8,
        "col_index": 71,
        "website": "https://www.windesheim.com/",
        "maps": "https://www.google.com/maps/place/Windesheim+University,+Zwolle,+Netherlands",
    },
    {
        "name": "Veleučilište u Šibeniku",
        "country": "Chorwacja",
        "city": "Šibenik",
        "qs_num": 9003,
        "qs": "b/r",
        "students": 1500,
        "temp": 9.8,
        "col_index": 48,
        "website": "https://www.vus.hr/en/",
        "maps": "https://www.google.com/maps/place/Polytechnic+of+Šibenik,+Croatia",
    },
]

# Styles
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
link_font = Font(name="Calibri", color="0563C1", underline="single", size=10)
data_font = Font(name="Calibri", size=10)
bold_font = Font(name="Calibri", bold=True, size=10)
alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
warm_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
cold_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
cheap_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
exp_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
thin_border = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

headers = [
    "#",
    "Uczelnia",
    "Kraj",
    "Miasto",
    "QS Ranking",
    "Liczba studentów",
    "Śr. temp. paź–lut (°C)",
    "Cost of Living Index\n(NYC=100, niżej=taniej)",
    "Strona uczelni",
    "Google Maps",
]

col_widths = [5, 44, 14, 18, 14, 17, 22, 26, 18, 18]


def build_sheet(ws, data, title_text):
    # Title
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = title_text
    title_cell.font = Font(name="Calibri", bold=True, size=13, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Legend row
    ws.merge_cells("A2:J2")
    legend = ws["A2"]
    legend.value = "b/r = brak w rankingu  |  Cost of Living: zielony ≤40 (tanio), czerwony ≥70 (drogo)  |  Temp: pomarańcz ≥10°C (ciepło), niebieski ≤0°C (zimno)"
    legend.font = Font(name="Calibri", italic=True, size=9, color="555555")
    legend.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    # Headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[3].height = 38

    # Data
    for i, uni in enumerate(data):
        row = i + 4
        fill = alt_fill if i % 2 == 0 else None

        def style_cell(cell, font=data_font, align_h="center", custom_fill=None):
            cell.font = font
            cell.alignment = Alignment(horizontal=align_h, vertical="center")
            cell.border = thin_border
            if custom_fill:
                cell.fill = custom_fill
            elif fill:
                cell.fill = fill

        # #
        c = ws.cell(row=row, column=1, value=i + 1)
        style_cell(c, font=bold_font)

        # University name
        c = ws.cell(row=row, column=2, value=uni["name"])
        style_cell(c, font=bold_font, align_h="left")

        # Country
        c = ws.cell(row=row, column=3, value=uni["country"])
        style_cell(c)

        # City
        c = ws.cell(row=row, column=4, value=uni["city"])
        style_cell(c)

        # QS Ranking
        c = ws.cell(row=row, column=5, value=uni["qs"])
        style_cell(c)

        # Students
        c = ws.cell(row=row, column=6, value=uni["students"])
        c.number_format = "#,##0"
        style_cell(c)

        # Temperature with conditional color
        temp = uni["temp"]
        t_fill = None
        if temp >= 10:
            t_fill = warm_fill
        elif temp <= 0:
            t_fill = cold_fill
        c = ws.cell(row=row, column=7, value=f"{temp:+.1f}°C")
        style_cell(c, custom_fill=t_fill or fill)

        # Cost of Living Index with conditional color
        col_idx_val = uni["col_index"]
        c_fill = None
        if col_idx_val <= 40:
            c_fill = cheap_fill
        elif col_idx_val >= 70:
            c_fill = exp_fill
        c = ws.cell(row=row, column=8, value=col_idx_val)
        style_cell(c, custom_fill=c_fill or fill)

        # Website
        c = ws.cell(row=row, column=9, value="Otwórz stronę")
        c.hyperlink = uni["website"]
        style_cell(c, font=link_font)

        # Google Maps
        c = ws.cell(row=row, column=10, value="Pokaż na mapie")
        c.hyperlink = uni["maps"]
        style_cell(c, font=link_font)

        ws.row_dimensions[row].height = 22

    # Column widths
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Freeze & filter
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:J{3 + len(data)}"


wb = openpyxl.Workbook()

# Sheet 1: sorted by QS ranking
ws1 = wb.active
ws1.title = "Wg rankingu QS"
by_ranking = sorted(universities, key=lambda u: u["qs_num"])
build_sheet(ws1, by_ranking, "Erasmus+ PŁ ∩ UŁ — sortowanie: QS World University Rankings")

# Sheet 2: sorted by university size (students, descending)
ws2 = wb.create_sheet("Wg wielkości uczelni")
by_size = sorted(universities, key=lambda u: u["students"], reverse=True)
build_sheet(ws2, by_size, "Erasmus+ PŁ ∩ UŁ — sortowanie: liczba studentów (malejąco)")

output_path = r"C:\Users\stani\OneDrive\Pulpit\Erasmus_PL_UL_wspolne.xlsx"
wb.save(output_path)
print(f"Saved to: {output_path}")
